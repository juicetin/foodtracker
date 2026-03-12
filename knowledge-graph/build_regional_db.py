#!/usr/bin/env python3
"""
Build regional nutrition SQLite databases from Australian (AFCD), UK (CoFID),
and French (CIQUAL) food composition data sources.

Converts each source's Excel format into the published nutrition pack schema
(foods, nutrients, food_nutrients, food_portions, foods_fts) for on-device use.

Usage:
    python build_regional_db.py --source afcd --input-dir /path/to/afcd_data --output afcd.db
    python build_regional_db.py --source cofid --input-dir /path/to/cofid_data --output cofid.db
    python build_regional_db.py --source ciqual --input-dir /path/to/ciqual_data --output ciqual.db

Sources:
    afcd    -- Australian Food Composition Database (foodstandards.gov.au)
    cofid   -- UK Composition of Foods Integrated Dataset (gov.uk)
    ciqual  -- French food composition database (anses.fr)
"""

import argparse
import logging
import os
import sqlite3
import sys
from pathlib import Path

import openpyxl

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

# Well-known nutrient IDs (matches USDA FDC numbering for schema compatibility)
NUTRIENT_IDS = {
    "Energy": 1008,
    "Protein": 1003,
    "Carbohydrate": 1005,
    "Fat": 1004,
    "Fibre": 1079,
    "Fiber": 1079,
}

# Nutrient definitions to insert into the nutrients table
STANDARD_NUTRIENTS = [
    {"id": 1008, "name": "Energy", "unit": "kcal", "nutrient_nbr": "208"},
    {"id": 1003, "name": "Protein", "unit": "g", "nutrient_nbr": "203"},
    {"id": 1005, "name": "Carbohydrate, by difference", "unit": "g", "nutrient_nbr": "205"},
    {"id": 1004, "name": "Total lipid (fat)", "unit": "g", "nutrient_nbr": "204"},
    {"id": 1079, "name": "Fiber, total dietary", "unit": "g", "nutrient_nbr": "291"},
]

SCHEMA_SQL = """
CREATE TABLE foods (
    fdc_id INTEGER PRIMARY KEY,
    description TEXT NOT NULL,
    food_category TEXT,
    data_type TEXT,
    publication_date TEXT
);

CREATE TABLE nutrients (
    id INTEGER PRIMARY KEY,
    name TEXT NOT NULL,
    unit TEXT NOT NULL,
    nutrient_nbr TEXT
);

CREATE TABLE food_nutrients (
    food_id INTEGER REFERENCES foods(fdc_id),
    nutrient_id INTEGER REFERENCES nutrients(id),
    amount REAL NOT NULL,
    PRIMARY KEY (food_id, nutrient_id)
);

CREATE TABLE food_portions (
    id INTEGER PRIMARY KEY,
    food_id INTEGER REFERENCES foods(fdc_id),
    portion_description TEXT,
    gram_weight REAL,
    modifier TEXT
);

-- FTS5 full-text search on food descriptions and categories
CREATE VIRTUAL TABLE foods_fts USING fts5(
    description,
    food_category,
    content=foods,
    content_rowid=fdc_id
);

-- Indexes for common queries
CREATE INDEX idx_foods_category ON foods(food_category);
CREATE INDEX idx_foods_data_type ON foods(data_type);
CREATE INDEX idx_food_nutrients_food ON food_nutrients(food_id);
CREATE INDEX idx_food_nutrients_nutrient ON food_nutrients(nutrient_id);
"""


def _load_afcd(input_dir: str):
    """
    Load AFCD data from Excel files.

    AFCD format:
    - food.xlsx: Food ID, Food Name, Food Category
    - nutrient.xlsx: Food ID, Nutrient ID, Nutrient Name, Unit, Amount
    """
    foods = []
    nutrients_data = []

    # Load foods
    food_path = os.path.join(input_dir, "food.xlsx")
    wb = openpyxl.load_workbook(food_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        food_id, name, category = row[0], row[1], row[2] if len(row) > 2 else None
        if food_id is not None and name is not None:
            foods.append({
                "fdc_id": int(food_id),
                "description": str(name),
                "food_category": str(category) if category else None,
            })
    wb.close()

    # Load nutrients
    nutrient_path = os.path.join(input_dir, "nutrient.xlsx")
    wb = openpyxl.load_workbook(nutrient_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        food_id, nutrient_id, nutrient_name, unit, amount = (
            row[0], row[1], row[2], row[3], row[4]
        )
        if food_id is not None and nutrient_id is not None and amount is not None:
            nutrients_data.append({
                "food_id": int(food_id),
                "nutrient_id": int(nutrient_id),
                "nutrient_name": str(nutrient_name),
                "unit": str(unit),
                "amount": float(amount),
            })
    wb.close()

    return foods, nutrients_data


def _load_cofid(input_dir: str):
    """
    Load CoFID data from a single Excel workbook.

    CoFID format:
    - cofid.xlsx with "Proximates" sheet:
      Food Code, Food Name, Group, Energy (kcal), Protein (g),
      Carbohydrate (g), Fat (g), Fibre (g)
    """
    foods = []
    nutrients_data = []

    cofid_path = os.path.join(input_dir, "cofid.xlsx")
    wb = openpyxl.load_workbook(cofid_path, read_only=True)
    ws = wb["Proximates"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    for idx, row in enumerate(rows, start=1):
        food_code, name, group = row[0], row[1], row[2]
        energy = row[3] if len(row) > 3 else None
        protein = row[4] if len(row) > 4 else None
        carbs = row[5] if len(row) > 5 else None
        fat = row[6] if len(row) > 6 else None
        fibre = row[7] if len(row) > 7 else None

        if food_code is not None and name is not None:
            fdc_id = idx  # Use row index as synthetic ID
            foods.append({
                "fdc_id": fdc_id,
                "description": str(name),
                "food_category": str(group) if group else None,
            })

            # Map nutrient columns
            nutrient_map = [
                (NUTRIENT_IDS["Energy"], "Energy", "kcal", energy),
                (NUTRIENT_IDS["Protein"], "Protein", "g", protein),
                (NUTRIENT_IDS["Carbohydrate"], "Carbohydrate", "g", carbs),
                (NUTRIENT_IDS["Fat"], "Fat", "g", fat),
                (NUTRIENT_IDS["Fibre"], "Fibre", "g", fibre),
            ]
            for nut_id, nut_name, unit, amount in nutrient_map:
                if amount is not None:
                    nutrients_data.append({
                        "food_id": fdc_id,
                        "nutrient_id": nut_id,
                        "nutrient_name": nut_name,
                        "unit": unit,
                        "amount": float(amount),
                    })

    wb.close()
    return foods, nutrients_data


def _load_ciqual(input_dir: str):
    """
    Load CIQUAL data from Excel with French column names.

    CIQUAL format:
    - ciqual.xlsx: Code, Nom, Groupe, Energie (kcal/100g),
      Proteines (g/100g), Glucides (g/100g), Lipides (g/100g)
    """
    foods = []
    nutrients_data = []

    ciqual_path = os.path.join(input_dir, "ciqual.xlsx")
    wb = openpyxl.load_workbook(ciqual_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    for idx, row in enumerate(rows, start=1):
        code, nom, groupe = row[0], row[1], row[2]
        energie = row[3] if len(row) > 3 else None
        proteines = row[4] if len(row) > 4 else None
        glucides = row[5] if len(row) > 5 else None
        lipides = row[6] if len(row) > 6 else None

        if code is not None and nom is not None:
            fdc_id = idx  # Use row index as synthetic ID
            foods.append({
                "fdc_id": fdc_id,
                "description": str(nom),
                "food_category": str(groupe) if groupe else None,
            })

            # Map French nutrient names to standard IDs
            nutrient_map = [
                (NUTRIENT_IDS["Energy"], "Energy", "kcal", energie),
                (NUTRIENT_IDS["Protein"], "Protein", "g", proteines),
                (NUTRIENT_IDS["Carbohydrate"], "Carbohydrate", "g", glucides),
                (NUTRIENT_IDS["Fat"], "Fat", "g", lipides),
            ]
            for nut_id, nut_name, unit, amount in nutrient_map:
                if amount is not None:
                    nutrients_data.append({
                        "food_id": fdc_id,
                        "nutrient_id": nut_id,
                        "nutrient_name": nut_name,
                        "unit": unit,
                        "amount": float(amount),
                    })

    wb.close()
    return foods, nutrients_data


LOADERS = {
    "afcd": _load_afcd,
    "cofid": _load_cofid,
    "ciqual": _load_ciqual,
}


def build_regional_db(source: str, input_dir: str, output: str):
    """
    Build a regional nutrition SQLite database from source Excel data.

    Args:
        source: Database source identifier ('afcd', 'cofid', 'ciqual')
        input_dir: Path to directory containing source Excel files
        output: Path for the output .db file
    """
    if source not in LOADERS:
        raise ValueError(f"Unknown source: {source}. Must be one of: {list(LOADERS.keys())}")

    logger.info(f"Building {source.upper()} regional pack from {input_dir}")
    logger.info(f"Output: {output}")

    # Remove existing output to start clean
    if os.path.exists(output):
        os.remove(output)

    # Load source data
    loader = LOADERS[source]
    foods, nutrients_data = loader(input_dir)

    if not foods:
        logger.error("No foods loaded from source data")
        sys.exit(1)

    logger.info(f"Loaded {len(foods)} foods, {len(nutrients_data)} nutrient values")

    # Create database and schema
    conn = sqlite3.connect(output)
    conn.execute("PRAGMA journal_mode = DELETE")
    conn.executescript(SCHEMA_SQL)

    # Insert standard nutrient definitions
    for nutrient in STANDARD_NUTRIENTS:
        conn.execute(
            "INSERT OR IGNORE INTO nutrients (id, name, unit, nutrient_nbr) VALUES (?, ?, ?, ?)",
            (nutrient["id"], nutrient["name"], nutrient["unit"], nutrient["nutrient_nbr"]),
        )

    # Also insert any extra nutrient definitions from the data
    seen_nutrient_ids = {n["id"] for n in STANDARD_NUTRIENTS}
    for nd in nutrients_data:
        if nd["nutrient_id"] not in seen_nutrient_ids:
            conn.execute(
                "INSERT OR IGNORE INTO nutrients (id, name, unit) VALUES (?, ?, ?)",
                (nd["nutrient_id"], nd["nutrient_name"], nd["unit"]),
            )
            seen_nutrient_ids.add(nd["nutrient_id"])

    logger.info(f"Inserted {len(seen_nutrient_ids)} nutrient definitions")

    # Insert foods with data_type set to source identifier
    for food in foods:
        conn.execute(
            "INSERT INTO foods (fdc_id, description, food_category, data_type) VALUES (?, ?, ?, ?)",
            (food["fdc_id"], food["description"], food["food_category"], source),
        )
    logger.info(f"Inserted {len(foods)} foods")

    # Insert food-nutrient links
    valid_food_ids = {f["fdc_id"] for f in foods}
    nutrient_count = 0
    for nd in nutrients_data:
        if nd["food_id"] in valid_food_ids:
            conn.execute(
                "INSERT OR IGNORE INTO food_nutrients (food_id, nutrient_id, amount) VALUES (?, ?, ?)",
                (nd["food_id"], nd["nutrient_id"], nd["amount"]),
            )
            nutrient_count += 1
    logger.info(f"Inserted {nutrient_count} food-nutrient links")

    # Populate FTS5 index
    conn.execute("""
        INSERT INTO foods_fts (rowid, description, food_category)
        SELECT fdc_id, description, food_category FROM foods
    """)
    logger.info("Built FTS5 search index")

    # Set schema version
    conn.execute("PRAGMA user_version = 1")

    # Commit all data before ANALYZE and VACUUM
    conn.commit()

    # Optimize
    conn.execute("ANALYZE")
    conn.execute("VACUUM")
    logger.info("Ran ANALYZE and VACUUM")

    conn.close()

    # Log summary
    file_size = os.path.getsize(output)
    size_kb = file_size / 1024
    logger.info(
        f"Build complete: {len(foods)} foods, {nutrient_count} nutrient links, "
        f"{size_kb:.1f} KB"
    )


def main():
    parser = argparse.ArgumentParser(
        description="Build regional nutrition SQLite database from Excel sources"
    )
    parser.add_argument(
        "--source",
        required=True,
        choices=["afcd", "cofid", "ciqual"],
        help="Regional data source: afcd (Australia), cofid (UK), ciqual (France)",
    )
    parser.add_argument(
        "--input-dir",
        required=True,
        help="Path to directory containing source Excel files",
    )
    parser.add_argument(
        "--output",
        required=True,
        help="Output .db file path",
    )
    args = parser.parse_args()

    build_regional_db(args.source, args.input_dir, args.output)


if __name__ == "__main__":
    main()
