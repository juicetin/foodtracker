#!/usr/bin/env python3
"""
Tests for regional nutrition database build pipeline.

Uses small fixture Excel data (5-10 foods each) to validate that AFCD, CoFID,
and CIQUAL source formats are correctly converted to the published nutrition
pack schema (foods, nutrients, food_nutrients, food_portions, foods_fts).
"""

import os
import sqlite3
import sys
from pathlib import Path

import openpyxl
import pytest

# Add parent directory to path so we can import the build module
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def create_afcd_fixtures(tmp_path: Path) -> Path:
    """Create small Excel fixtures mimicking AFCD format from foodstandards.gov.au."""
    afcd_dir = tmp_path / "afcd_input"
    afcd_dir.mkdir()

    # food.xlsx -- AFCD food items
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Food"
    ws.append(["Food ID", "Food Name", "Food Category"])
    ws.append([1001, "Chicken breast, raw", "Poultry"])
    ws.append([1002, "Lamb chop, grilled", "Meat"])
    ws.append([1003, "Vegemite", "Spreads"])
    ws.append([1004, "Tim Tam, chocolate", "Biscuits"])
    ws.append([1005, "Kangaroo steak, raw", "Game Meats"])
    wb.save(afcd_dir / "food.xlsx")

    # nutrient.xlsx -- AFCD nutrient definitions and values
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Nutrient"
    ws2.append(["Food ID", "Nutrient ID", "Nutrient Name", "Unit", "Amount"])
    # Chicken breast nutrients
    ws2.append([1001, 1008, "Energy", "kcal", 110])
    ws2.append([1001, 1003, "Protein", "g", 23.1])
    ws2.append([1001, 1005, "Carbohydrate", "g", 0.0])
    ws2.append([1001, 1004, "Fat", "g", 1.2])
    # Lamb chop nutrients
    ws2.append([1002, 1008, "Energy", "kcal", 250])
    ws2.append([1002, 1003, "Protein", "g", 25.5])
    ws2.append([1002, 1005, "Carbohydrate", "g", 0.0])
    ws2.append([1002, 1004, "Fat", "g", 16.0])
    # Vegemite nutrients
    ws2.append([1003, 1008, "Energy", "kcal", 174])
    ws2.append([1003, 1003, "Protein", "g", 25.4])
    ws2.append([1003, 1005, "Carbohydrate", "g", 14.6])
    ws2.append([1003, 1004, "Fat", "g", 0.5])
    # Tim Tam nutrients
    ws2.append([1004, 1008, "Energy", "kcal", 486])
    ws2.append([1004, 1003, "Protein", "g", 5.5])
    ws2.append([1004, 1005, "Carbohydrate", "g", 62.0])
    ws2.append([1004, 1004, "Fat", "g", 24.0])
    # Kangaroo steak nutrients
    ws2.append([1005, 1008, "Energy", "kcal", 98])
    ws2.append([1005, 1003, "Protein", "g", 22.0])
    ws2.append([1005, 1005, "Carbohydrate", "g", 0.0])
    ws2.append([1005, 1004, "Fat", "g", 0.9])
    wb2.save(afcd_dir / "nutrient.xlsx")

    return afcd_dir


def create_cofid_fixtures(tmp_path: Path) -> Path:
    """Create small Excel fixtures mimicking CoFID format from gov.uk."""
    cofid_dir = tmp_path / "cofid_input"
    cofid_dir.mkdir()

    # CoFID is a single workbook with multiple sheets
    wb = openpyxl.Workbook()

    # "Proximates" sheet -- main nutrient data
    ws = wb.active
    ws.title = "Proximates"
    ws.append(["Food Code", "Food Name", "Group", "Energy (kcal)", "Protein (g)",
               "Carbohydrate (g)", "Fat (g)", "Fibre (g)"])
    ws.append(["C001", "Bangers and mash", "Meat Dishes", 150, 8.5, 15.0, 6.2, 1.8])
    ws.append(["C002", "Fish and chips", "Fish Dishes", 247, 12.0, 28.0, 10.5, 2.0])
    ws.append(["C003", "Cornish pasty", "Meat Dishes", 267, 7.5, 24.5, 15.5, 1.5])
    ws.append(["C004", "Baked beans on toast", "Cereal Products", 168, 8.0, 28.0, 2.5, 5.5])
    ws.append(["C005", "Yorkshire pudding", "Cereal Products", 208, 7.3, 27.0, 8.2, 0.9])
    ws.append(["C006", "Scones, plain", "Cereal Products", 362, 7.2, 53.0, 14.0, 1.8])
    wb.save(cofid_dir / "cofid.xlsx")

    return cofid_dir


def create_ciqual_fixtures(tmp_path: Path) -> Path:
    """Create small Excel fixtures mimicking CIQUAL format from anses.fr."""
    ciqual_dir = tmp_path / "ciqual_input"
    ciqual_dir.mkdir()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CIQUAL"
    # CIQUAL uses French column names
    ws.append(["Code", "Nom", "Groupe", "Energie (kcal/100g)",
               "Proteines (g/100g)", "Glucides (g/100g)", "Lipides (g/100g)"])
    ws.append(["CQ001", "Croissant au beurre", "Patisseries", 406, 8.2, 45.0, 21.0])
    ws.append(["CQ002", "Baguette tradition", "Pains", 285, 9.0, 56.0, 1.5])
    ws.append(["CQ003", "Camembert", "Fromages", 299, 20.5, 0.5, 24.0])
    ws.append(["CQ004", "Ratatouille", "Plats composes", 55, 1.5, 6.0, 2.5])
    ws.append(["CQ005", "Quiche lorraine", "Plats composes", 230, 10.0, 18.0, 13.5])
    wb.save(ciqual_dir / "ciqual.xlsx")

    return ciqual_dir


@pytest.fixture
def afcd_dir(tmp_path):
    return create_afcd_fixtures(tmp_path)


@pytest.fixture
def cofid_dir(tmp_path):
    return create_cofid_fixtures(tmp_path)


@pytest.fixture
def ciqual_dir(tmp_path):
    return create_ciqual_fixtures(tmp_path)


@pytest.fixture
def output_db(tmp_path):
    return str(tmp_path / "regional_nutrition.db")


class TestAFCDBuild:
    """Test AFCD (Australian Food Composition Database) build pipeline."""

    def test_afcd_converts_to_valid_sqlite(self, afcd_dir, output_db):
        """AFCD Excel data converts to valid SQLite matching nutrition pack schema."""
        from build_regional_db import build_regional_db

        build_regional_db(source="afcd", input_dir=str(afcd_dir), output=output_db)

        assert os.path.exists(output_db)
        conn = sqlite3.connect(output_db)
        cursor = conn.cursor()

        # Check foods table exists and has data
        cursor.execute("SELECT COUNT(*) FROM foods")
        count = cursor.fetchone()[0]
        assert count == 5

        # Check a specific food
        cursor.execute("SELECT description FROM foods WHERE fdc_id = 1001")
        row = cursor.fetchone()
        assert row is not None
        assert "Chicken breast" in row[0]

        conn.close()

    def test_afcd_has_correct_data_type(self, afcd_dir, output_db):
        """AFCD output database has data_type set to 'afcd'."""
        from build_regional_db import build_regional_db

        build_regional_db(source="afcd", input_dir=str(afcd_dir), output=output_db)

        conn = sqlite3.connect(output_db)
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT data_type FROM foods")
        types = [row[0] for row in cursor.fetchall()]
        assert types == ["afcd"]
        conn.close()

    def test_afcd_fts5_search_works(self, afcd_dir, output_db):
        """FTS5 index works in AFCD regional database."""
        from build_regional_db import build_regional_db

        build_regional_db(source="afcd", input_dir=str(afcd_dir), output=output_db)

        conn = sqlite3.connect(output_db)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT f.fdc_id, f.description
            FROM foods_fts fts
            JOIN foods f ON f.fdc_id = fts.rowid
            WHERE foods_fts MATCH 'chicken'
        """)
        results = cursor.fetchall()
        assert len(results) >= 1
        assert any("Chicken" in r[1] for r in results)
        conn.close()


class TestCoFIDBuild:
    """Test CoFID (UK Composition of Foods Integrated Dataset) build pipeline."""

    def test_cofid_converts_to_valid_sqlite(self, cofid_dir, output_db):
        """CoFID Excel data converts to valid SQLite matching nutrition pack schema."""
        from build_regional_db import build_regional_db

        build_regional_db(source="cofid", input_dir=str(cofid_dir), output=output_db)

        assert os.path.exists(output_db)
        conn = sqlite3.connect(output_db)
        cursor = conn.cursor()

        cursor.execute("SELECT COUNT(*) FROM foods")
        count = cursor.fetchone()[0]
        assert count == 6

        cursor.execute("SELECT description FROM foods WHERE fdc_id = 1")
        row = cursor.fetchone()
        assert row is not None
        assert "Bangers" in row[0]

        conn.close()

    def test_cofid_has_correct_data_type(self, cofid_dir, output_db):
        """CoFID output database has data_type set to 'cofid'."""
        from build_regional_db import build_regional_db

        build_regional_db(source="cofid", input_dir=str(cofid_dir), output=output_db)

        conn = sqlite3.connect(output_db)
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT data_type FROM foods")
        types = [row[0] for row in cursor.fetchall()]
        assert types == ["cofid"]
        conn.close()

    def test_cofid_fts5_search_works(self, cofid_dir, output_db):
        """FTS5 index works in CoFID regional database."""
        from build_regional_db import build_regional_db

        build_regional_db(source="cofid", input_dir=str(cofid_dir), output=output_db)

        conn = sqlite3.connect(output_db)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT f.fdc_id, f.description
            FROM foods_fts fts
            JOIN foods f ON f.fdc_id = fts.rowid
            WHERE foods_fts MATCH 'fish'
        """)
        results = cursor.fetchall()
        assert len(results) >= 1
        conn.close()


class TestCIQUALBuild:
    """Test CIQUAL (French food composition database) build pipeline."""

    def test_ciqual_converts_to_valid_sqlite(self, ciqual_dir, output_db):
        """CIQUAL Excel data converts to valid SQLite matching nutrition pack schema."""
        from build_regional_db import build_regional_db

        build_regional_db(source="ciqual", input_dir=str(ciqual_dir), output=output_db)

        assert os.path.exists(output_db)
        conn = sqlite3.connect(output_db)
        cursor = conn.cursor()

        cursor.execute("SELECT COUNT(*) FROM foods")
        count = cursor.fetchone()[0]
        assert count == 5

        cursor.execute("SELECT description FROM foods WHERE fdc_id = 1")
        row = cursor.fetchone()
        assert row is not None
        assert "Croissant" in row[0]

        conn.close()

    def test_ciqual_has_correct_data_type(self, ciqual_dir, output_db):
        """CIQUAL output database has data_type set to 'ciqual'."""
        from build_regional_db import build_regional_db

        build_regional_db(source="ciqual", input_dir=str(ciqual_dir), output=output_db)

        conn = sqlite3.connect(output_db)
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT data_type FROM foods")
        types = [row[0] for row in cursor.fetchall()]
        assert types == ["ciqual"]
        conn.close()

    def test_ciqual_fts5_search_works(self, ciqual_dir, output_db):
        """FTS5 index works in CIQUAL regional database."""
        from build_regional_db import build_regional_db

        build_regional_db(source="ciqual", input_dir=str(ciqual_dir), output=output_db)

        conn = sqlite3.connect(output_db)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT f.fdc_id, f.description
            FROM foods_fts fts
            JOIN foods f ON f.fdc_id = fts.rowid
            WHERE foods_fts MATCH 'camembert'
        """)
        results = cursor.fetchall()
        assert len(results) >= 1
        conn.close()


class TestRegionalBuildCommon:
    """Test common behavior across all regional database builds."""

    def test_nutrients_table_populated(self, afcd_dir, output_db):
        """Nutrients table has nutrient definitions in regional DB."""
        from build_regional_db import build_regional_db

        build_regional_db(source="afcd", input_dir=str(afcd_dir), output=output_db)

        conn = sqlite3.connect(output_db)
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM nutrients")
        count = cursor.fetchone()[0]
        assert count > 0
        conn.close()

    def test_food_nutrients_linked(self, afcd_dir, output_db):
        """food_nutrients table links foods to nutrients with amounts."""
        from build_regional_db import build_regional_db

        build_regional_db(source="afcd", input_dir=str(afcd_dir), output=output_db)

        conn = sqlite3.connect(output_db)
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM food_nutrients")
        count = cursor.fetchone()[0]
        assert count > 0

        # Check specific food has energy
        cursor.execute("""
            SELECT fn.amount
            FROM food_nutrients fn
            JOIN nutrients n ON fn.nutrient_id = n.id
            WHERE fn.food_id = 1001 AND n.name = 'Energy'
        """)
        row = cursor.fetchone()
        assert row is not None
        assert row[0] == 110  # AFCD chicken breast energy
        conn.close()

    def test_pragma_user_version_set(self, afcd_dir, output_db):
        """PRAGMA user_version is set in regional DB."""
        from build_regional_db import build_regional_db

        build_regional_db(source="afcd", input_dir=str(afcd_dir), output=output_db)

        conn = sqlite3.connect(output_db)
        cursor = conn.cursor()
        cursor.execute("PRAGMA user_version")
        version = cursor.fetchone()[0]
        assert version == 1
        conn.close()

    def test_schema_matches_published_spec(self, cofid_dir, output_db):
        """Output database tables match the published nutrition pack schema exactly."""
        from build_regional_db import build_regional_db

        build_regional_db(source="cofid", input_dir=str(cofid_dir), output=output_db)

        conn = sqlite3.connect(output_db)
        cursor = conn.cursor()

        # Check required tables exist
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
        tables = [row[0] for row in cursor.fetchall()]
        for required in ["foods", "nutrients", "food_nutrients", "food_portions"]:
            assert required in tables, f"Missing table: {required}"

        # Check FTS5 virtual table
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='foods_fts'")
        assert cursor.fetchone() is not None, "Missing FTS5 virtual table: foods_fts"

        conn.close()
